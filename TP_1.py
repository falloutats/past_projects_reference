def main():

    city, state = data_splitter(wb.sheetnames[31:165])

    counter1 = 0
    counter2 = 0
    not_in_df = list()

    for i in range(len(city)):
        c_ity = city[i]
        s_tate = state[i]

        data = fetch_data(filer_by_state=s_tate, filter_by_city=c_ity, info=False)
        if not data.empty:
            print(data)
            counter1 += 1
        else:
            not_in_df.append(city)
            counter2 += 1

    print(f'Cities in the df: {counter1} cities not in the df {counter2} {len(not_in_df)}')


def fetch_data(filer_by_state=None, filter_by_city=None, info=False):
    """
    :param filer_by_state: filter df by state
    :param filter_by_city: filter df by city
    :param info: display output T OR F
    :return: return df after applying filter
    """
    target = df[(df['State'] == filer_by_state) & (df['City'] == filter_by_city)]

    if info:
        print(f'Shape of df: {df.shape}')

    return target[target.columns[1:7]]


def match(filter_data,index=31,  reveal=True):

    """
    :param index: sheet index within a workbook
    :param filter_data: data that has been filtered by fetch_data
    :param reveal: print info of changes being applied
    :return: number of times the loop was executed
    """
    print(f'index working on {index}')

    count = 0
    check_col = 29
    change_col = 32

    wb[wb.sheetnames[index]].cell(row=7, column=32).value = 'Q1 2023'

    print(wb[wb.sheetnames[index]])

    for row in range(8, 20):

        item_name = wb[wb.sheetnames[index]].cell(row=row, column=check_col).value
        mask = filter_data['Product Name'] == item_name
        original_price = float(filter_data[mask]['Original Price'])
        discounted_price = float(filter_data[mask]['Discounted Price'])
        count += 1

        if reveal:
            print(
                f'Item Name: {item_name}| Original Price: {original_price} {type(original_price)}|  Discount Price: {discounted_price}{type(discounted_price)}')

        if discounted_price > 0.0:
            wb[wb.sheetnames[index]].cell(row=row, column=change_col).value = discounted_price
            print(f'Value changed at row: {row} to {discounted_price}')

        else:
            wb[wb.sheetnames[index]].cell(row=row, column=change_col).value = original_price
            print(f'Value changed at row: {row} to {original_price}')

    return f'Completed and ran a total of {count} times'


def data_splitter(sheet_names):

    """
    :param sheet_names: list of names of sheets within a workbook
    :return: items as city, state list with no null values
    """
    city_name = list()
    state_name = list()
    for item in sheet_names:
        a, b, c = item.partition(', ')

        if c == '':

            a, b, c = item.partition(' ')

            state_name.append(c)
            city_name.append(a)
        else:

            state_name.append(c)
            city_name.append(a)

    return city_name, state_name


if __name__ == '__main__':
    import openpyxl
    import pandas as pd
    import time

    file_with_data = '/Users/fallout/PycharmProjects/union_rates_time_series/data.xlsx'
    target_file = '/Users/fallout/Downloads/location.xlsx'

    ####################################################################################################################

    tick = time.time()
    df = pd.read_excel(file_with_data)
    wb = openpyxl.load_workbook(target_file)
    tock = time.time()
    city_state_index = 0
    print(f'{tock - tick} seconds, to open the file')

    city, state = data_splitter(wb.sheetnames[31:165])

    for i in range(31,165):

        data = fetch_data(filer_by_state=state[city_state_index], filter_by_city=city[city_state_index])

        if data.empty:
            ...
        else:
            match(filter_data=data ,index=i)

        city_state_index+= 1


    wb.save('test_final.xlsx')


